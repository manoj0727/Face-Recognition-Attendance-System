import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
import os
from typing import Dict, List, Optional

class EnhancedExcelExporter:
    def __init__(self, db_manager):
        self.db = db_manager
        self.export_dir = 'attendance_records'
        os.makedirs(self.export_dir, exist_ok=True)
        
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.late_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
    def export_classroom_attendance(self, classroom_id: str, date: str = None, 
                                   include_analytics: bool = True) -> str:
        """Export attendance for a specific classroom"""
        wb = Workbook()
        
        if date:
            self._add_daily_attendance_sheet(wb, classroom_id, date)
        else:
            self._add_complete_attendance_sheet(wb, classroom_id)
        
        self._add_student_summary_sheet(wb, classroom_id)
        
        if include_analytics:
            self._add_analytics_sheet(wb, classroom_id)
            self._add_charts_sheet(wb, classroom_id)
        
        self._add_metadata_sheet(wb, classroom_id)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        filename = f"{self.export_dir}/classroom_{classroom_id}_attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
    
    def _add_daily_attendance_sheet(self, wb: Workbook, classroom_id: str, date: str):
        """Add daily attendance sheet"""
        ws = wb.create_sheet(f"Attendance_{date}")
        
        attendance_records = self.db.get_classroom_attendance(classroom_id, date)
        all_students = self.db.get_classroom_students(classroom_id)
        
        headers = ['S.No', 'Student ID', 'Name', 'Department', 'Year', 'Time', 'Status', 'Confidence']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        attended_students = {record[0]: record for record in attendance_records}
        
        row_num = 2
        for idx, student in enumerate(all_students, 1):
            student_id = student[0]
            
            if student_id in attended_students:
                record = attended_students[student_id]
                time = record[2]
                status = record[3]
                confidence = f"{record[4]:.2f}" if record[4] else "N/A"
            else:
                time = "-"
                status = "A"
                confidence = "-"
            
            row_data = [idx, student_id, student[1], student[3], student[4], time, status, confidence]
            ws.append(row_data)
            
            status_cell = ws.cell(row=row_num, column=7)
            if status == 'P':
                status_cell.fill = self.present_fill
            elif status == 'A':
                status_cell.fill = self.absent_fill
            
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
            
            row_num += 1
        
        ws.append([])
        ws.append(['Summary'])
        ws.append(['Total Students', len(all_students)])
        ws.append(['Present', sum(1 for r in attendance_records if r[3] == 'P')])
        ws.append(['Absent', len(all_students) - sum(1 for r in attendance_records if r[3] == 'P')])
        ws.append(['Attendance %', f"{(sum(1 for r in attendance_records if r[3] == 'P') / len(all_students) * 100):.2f}%" if all_students else "0%"])
        
        self._adjust_column_widths(ws)
    
    def _add_complete_attendance_sheet(self, wb: Workbook, classroom_id: str):
        """Add complete attendance history sheet"""
        ws = wb.create_sheet("Complete_Attendance")
        
        all_records = self.db.get_classroom_attendance(classroom_id)
        
        if not all_records:
            ws.append(['No attendance records found'])
            return
        
        headers = ['S.No', 'Date', 'Student ID', 'Name', 'Time', 'Status', 'Confidence']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, record in enumerate(all_records, 1):
            row_data = [idx, record[2], record[0], record[1], record[3], record[4], 
                       f"{record[5]:.2f}" if record[5] else "N/A"]
            ws.append(row_data)
            
            status_cell = ws.cell(row=idx + 1, column=6)
            if record[4] == 'P':
                status_cell.fill = self.present_fill
            elif record[4] == 'A':
                status_cell.fill = self.absent_fill
        
        self._adjust_column_widths(ws)
    
    def _add_student_summary_sheet(self, wb: Workbook, classroom_id: str):
        """Add individual student summary sheet"""
        ws = wb.create_sheet("Student_Summary")
        
        statistics = self.db.get_attendance_statistics(classroom_id)
        
        headers = ['S.No', 'Student ID', 'Name', 'Present Days', 'Absent Days', 
                  'Total Days', 'Attendance %', 'Status']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, stat in enumerate(statistics, 1):
            attendance_pct = stat[5] if stat[5] else 0
            
            if attendance_pct >= 75:
                status = "Good"
                status_color = self.present_fill
            elif attendance_pct >= 60:
                status = "Warning"
                status_color = self.late_fill
            else:
                status = "Critical"
                status_color = self.absent_fill
            
            row_data = [idx, stat[0], stat[1], stat[2], stat[3], stat[4], f"{attendance_pct:.2f}%", status]
            ws.append(row_data)
            
            ws.cell(row=idx + 1, column=7).fill = status_color
            ws.cell(row=idx + 1, column=8).fill = status_color
        
        self._adjust_column_widths(ws)
    
    def _add_analytics_sheet(self, wb: Workbook, classroom_id: str):
        """Add analytics and insights sheet"""
        ws = wb.create_sheet("Analytics")
        
        statistics = self.db.get_attendance_statistics(classroom_id)
        
        ws.append(['Attendance Analytics Report'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        
        ws.append(['Key Metrics'])
        ws['A3'].font = Font(bold=True, size=12)
        
        total_students = len(statistics)
        if statistics:
            avg_attendance = sum(s[5] for s in statistics if s[5]) / len([s for s in statistics if s[5]])
            students_above_90 = sum(1 for s in statistics if s[5] and s[5] > 90)
            students_75_90 = sum(1 for s in statistics if s[5] and 75 <= s[5] <= 90)
            students_below_75 = sum(1 for s in statistics if s[5] and s[5] < 75)
        else:
            avg_attendance = 0
            students_above_90 = 0
            students_75_90 = 0
            students_below_75 = 0
        
        metrics = [
            ['Total Students', total_students],
            ['Average Attendance %', f"{avg_attendance:.2f}%"],
            ['Students with >90% Attendance', students_above_90],
            ['Students with 75-90% Attendance', students_75_90],
            ['Students with <75% Attendance', students_below_75],
        ]
        
        for metric in metrics:
            ws.append(metric)
        
        ws.append([])
        ws.append(['Top Performers (>90% Attendance)'])
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        
        top_performers = [s for s in statistics if s[5] and s[5] > 90]
        for performer in sorted(top_performers, key=lambda x: x[5], reverse=True)[:10]:
            ws.append([performer[1], f"{performer[5]:.2f}%"])
        
        ws.append([])
        ws.append(['Students Requiring Attention (<75% Attendance)'])
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        
        low_performers = [s for s in statistics if s[5] and s[5] < 75]
        for performer in sorted(low_performers, key=lambda x: x[5]):
            ws.append([performer[1], f"{performer[5]:.2f}%"])
            ws.cell(row=ws.max_row, column=2).fill = self.absent_fill
        
        self._adjust_column_widths(ws)
    
    def _add_charts_sheet(self, wb: Workbook, classroom_id: str):
        """Add visual charts sheet"""
        ws = wb.create_sheet("Charts")
        
        statistics = self.db.get_attendance_statistics(classroom_id)
        
        data = []
        for stat in statistics:
            data.append([stat[1], stat[2], stat[3]])
        
        df = pd.DataFrame(data, columns=['Student', 'Present', 'Absent'])
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Attendance Overview"
        chart.y_axis.title = 'Days'
        chart.x_axis.title = 'Students'
        
        data = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1, max_col=3)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 20
        
        ws.add_chart(chart, "F2")
        
        if statistics:
            attendance_ranges = {
                '>90%': sum(1 for s in statistics if s[5] and s[5] > 90),
                '75-90%': sum(1 for s in statistics if s[5] and 75 <= s[5] <= 90),
                '60-75%': sum(1 for s in statistics if s[5] and 60 <= s[5] < 75),
                '<60%': sum(1 for s in statistics if s[5] and s[5] < 60)
            }
            
            ws.cell(row=2, column=8, value="Attendance Distribution")
            row = 3
            for range_label, count in attendance_ranges.items():
                ws.cell(row=row, column=8, value=range_label)
                ws.cell(row=row, column=9, value=count)
                row += 1
            
            pie = PieChart()
            pie.title = "Attendance Distribution"
            data = Reference(ws, min_col=9, min_row=3, max_row=6)
            labels = Reference(ws, min_col=8, min_row=3, max_row=6)
            pie.add_data(data)
            pie.set_categories(labels)
            pie.height = 10
            pie.width = 10
            
            ws.add_chart(pie, "K2")
    
    def _add_metadata_sheet(self, wb: Workbook, classroom_id: str):
        """Add metadata and export information sheet"""
        ws = wb.create_sheet("Info")
        
        classrooms = self.db.get_all_classrooms()
        classroom_info = next((c for c in classrooms if c[0] == classroom_id), None)
        
        ws.append(['Export Information'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        
        info = [
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Classroom ID', classroom_id],
            ['Classroom Name', classroom_info[1] if classroom_info else 'N/A'],
            ['Department', classroom_info[2] if classroom_info else 'N/A'],
            ['Semester', classroom_info[3] if classroom_info else 'N/A'],
            ['Academic Year', classroom_info[4] if classroom_info else 'N/A'],
            ['Instructor', classroom_info[5] if classroom_info else 'N/A'],
        ]
        
        for item in info:
            ws.append(item)
        
        self._adjust_column_widths(ws)
    
    def _adjust_column_widths(self, ws):
        """Automatically adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_batch_reports(self, classroom_ids: List[str], export_type: str = 'individual') -> List[str]:
        """Export reports for multiple classrooms"""
        exported_files = []
        
        for classroom_id in classroom_ids:
            try:
                filename = self.export_classroom_attendance(classroom_id)
                exported_files.append(filename)
            except Exception as e:
                print(f"Error exporting classroom {classroom_id}: {e}")
        
        return exported_files