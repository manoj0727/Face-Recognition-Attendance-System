import pandas as pd
from datetime import datetime
import os
from database_manager import DatabaseManager

class ExcelExporter:
    def __init__(self):
        self.db_manager = DatabaseManager()
        self.export_dir = 'attendance_records'
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_daily_attendance(self, date=None):
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
        
        records = self.db_manager.get_attendance_by_date(date)
        
        if not records:
            return None, "No attendance records found for this date"
        
        df = pd.DataFrame(records, columns=['Student ID', 'Name', 'Time', 'Status'])
        
        filename = f"{self.export_dir}/attendance_{date}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import PatternFill, Font
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            present_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            absent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=4):
                for cell in row:
                    if cell.value == 'P':
                        cell.fill = present_fill
                    elif cell.value == 'A':
                        cell.fill = absent_fill
        
        return filename, f"Attendance exported successfully to {filename}"
    
    def export_complete_attendance(self):
        records = self.db_manager.get_all_attendance()
        
        if not records:
            return None, "No attendance records found"
        
        df = pd.DataFrame(records, columns=['Student ID', 'Name', 'Date', 'Time', 'Status'])
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{self.export_dir}/complete_attendance_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All Attendance', index=False)
            
            summary_data = []
            student_ids = df['Student ID'].unique()
            
            for student_id in student_ids:
                student_records = df[df['Student ID'] == student_id]
                name = student_records['Name'].iloc[0]
                total_days = len(student_records)
                present_days = len(student_records[student_records['Status'] == 'P'])
                absent_days = len(student_records[student_records['Status'] == 'A'])
                attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
                
                summary_data.append({
                    'Student ID': student_id,
                    'Name': name,
                    'Total Days': total_days,
                    'Present': present_days,
                    'Absent': absent_days,
                    'Attendance %': f"{attendance_percentage:.2f}%"
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            workbook = writer.book
            
            for sheet_name in ['All Attendance', 'Summary']:
                worksheet = writer.sheets[sheet_name]
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                from openpyxl.styles import PatternFill, Font
                
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
        
        return filename, f"Complete attendance exported successfully to {filename}"
    
    def export_student_report(self, student_id):
        all_records = self.db_manager.get_all_attendance()
        student_records = [r for r in all_records if r[0] == student_id]
        
        if not student_records:
            return None, "No records found for this student"
        
        df = pd.DataFrame(student_records, columns=['Student ID', 'Name', 'Date', 'Time', 'Status'])
        
        student_name = df['Name'].iloc[0].replace(' ', '_')
        filename = f"{self.export_dir}/student_report_{student_id}_{student_name}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance Details', index=False)
            
            total_days = len(df)
            present_days = len(df[df['Status'] == 'P'])
            absent_days = len(df[df['Status'] == 'A'])
            attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
            
            summary_data = {
                'Metric': ['Total Days', 'Present Days', 'Absent Days', 'Attendance Percentage'],
                'Value': [total_days, present_days, absent_days, f"{attendance_percentage:.2f}%"]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            workbook = writer.book
            
            for sheet_name in ['Attendance Details', 'Summary']:
                worksheet = writer.sheets[sheet_name]
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filename, f"Student report exported successfully to {filename}"